"""Generates 04_Zava_Divisional_Variance_FY2025.xlsx — 4 sheets focused on FY25 variance.

Sheets:
  1. Variance Summary — 11 div × 12 cols (bud/act/var/var%/status RAG)
  2. Driver Tree — 5 worst divisions × 6 driver categories × 4 sub-drivers
  3. Monthly Variance Trend — 11 div × 12 months × 4 metrics
  4. Recovery Levers — 30+ specific actions × 6 attributes
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from gen_ref_files.common import DIVISIONS, FY25, fy25_eb, status_color

NAVY_FILL = PatternFill('solid', fgColor='1F4E79')
SUB_FILL = PatternFill('solid', fgColor='2E75B6')
TOTAL_FILL = PatternFill('solid', fgColor='D9E1F2')
RED_FILL = PatternFill('solid', fgColor='C00000')
AMBER_FILL = PatternFill('solid', fgColor='FFC000')
GREEN_FILL = PatternFill('solid', fgColor='00B050')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Aptos')
SUB_FONT = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
BODY_FONT = Font(size=10, name='Aptos')
TOTAL_FONT = Font(bold=True, size=10, name='Aptos')
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0;[Red](#,##0);-'
PCT_FMT = '0.0%;[Red]-0.0%;-'


def style_h(c, fill=NAVY_FILL, font=HEADER_FONT):
    c.fill = fill
    c.font = font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER


def status_pill(cell, status):
    if status == 'red':
        cell.fill = RED_FILL
        cell.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
    elif status == 'amber':
        cell.fill = AMBER_FILL
        cell.font = Font(bold=True, color='000000', size=10, name='Aptos')
    elif status == 'green':
        cell.fill = GREEN_FILL
        cell.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')


# ── Sheet 1: Variance Summary ──────────────────────────────────────────────
def s1_summary(ws):
    ws.title = 'Variance Summary'
    ws['A1'] = 'FY2025 Divisional Variance Summary — Budget vs Actual (MYR M)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:L1')
    ws['A2'] = 'Source: Group FP&A | Reporting period FY2025 actuals through Dec | RAG status thresholds: Green ≤2% miss, Amber 2-10%, Red >10%'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:L2')

    headers = ['Division', 'Country', 'Bud Rev', 'Act Rev', 'Rev Var', 'Rev Var %',
               'Bud EBITDA', 'Act EBITDA', 'EBITDA Var', 'EBITDA Var %', 'Status', 'Comment']
    for i, h in enumerate(headers, 1):
        style_h(ws.cell(row=4, column=i, value=h))

    row = 5
    COMMENTS = {
        'chem': 'China demand softness; mix shift to commodity grade',
        'agri': 'CPO price decline 18%; ID export levy increase',
        'mfg': 'Auto OEM volume softness; Penang plant downtime',
        'fs': 'Margin lending growth; underwriting fees beat',
        'prop': 'Mid-Valley delay; Penang Tower writedown MYR 320M',
        'bpo': 'New US client wins; AI productivity uplift',
        'trade': 'Hindustan Bulk counterparty event MYR 65M',
        'pharma': 'Generics tender wins ID; API cost manageable',
        'hc': 'Patient volume −11%; Subang at 58% utilisation',
        'retail': 'Footfall −14%; inventory writedown MYR 110M',
        'treas': 'Investment portfolio mark-down; FX translation',
    }
    for did, dname, country, _, _ in DIVISIONS:
        b_rev, a_rev, b_pct, a_pct, status = FY25[did]
        b_eb = fy25_eb(did, 'bud_eb')
        a_eb = fy25_eb(did, 'act_eb')
        rev_var = a_rev - b_rev
        rev_var_pct = rev_var / b_rev if b_rev else 0
        eb_var = a_eb - b_eb
        eb_var_pct = eb_var / b_eb if b_eb else 0

        ws.cell(row=row, column=1, value=dname).font = BODY_FONT
        ws.cell(row=row, column=2, value=country).font = BODY_FONT
        ws.cell(row=row, column=3, value=b_rev).number_format = NUM_FMT
        ws.cell(row=row, column=4, value=a_rev).number_format = NUM_FMT
        c = ws.cell(row=row, column=5, value=rev_var); c.number_format = NUM_FMT
        if rev_var < 0:
            c.font = Font(color='C00000', size=10, name='Aptos')
        c = ws.cell(row=row, column=6, value=rev_var_pct); c.number_format = PCT_FMT
        if rev_var_pct < 0:
            c.font = Font(color='C00000', size=10, name='Aptos')
        ws.cell(row=row, column=7, value=round(b_eb, 1)).number_format = NUM_FMT
        ws.cell(row=row, column=8, value=round(a_eb, 1)).number_format = NUM_FMT
        c = ws.cell(row=row, column=9, value=round(eb_var, 1)); c.number_format = NUM_FMT
        if eb_var < 0:
            c.font = Font(color='C00000', size=10, name='Aptos', bold=True)
        c = ws.cell(row=row, column=10, value=eb_var_pct); c.number_format = PCT_FMT
        if eb_var_pct < 0:
            c.font = Font(color='C00000', size=10, name='Aptos', bold=True)

        sc = ws.cell(row=row, column=11, value=status.upper())
        status_pill(sc, status)
        sc.alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=12, value=COMMENTS[did]).font = BODY_FONT
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    # Group total
    ws.cell(row=row, column=1, value='ZAVA GROUP CONSOLIDATED').font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    ws.cell(row=row, column=2, value='—').font = TOTAL_FONT
    ws.cell(row=row, column=2).fill = TOTAL_FILL
    bud_rev_total = sum(FY25[d][0] for d in [r[0] for r in DIVISIONS])
    act_rev_total = sum(FY25[d][1] for d in [r[0] for r in DIVISIONS])
    bud_eb_total = sum(fy25_eb(d, 'bud_eb') for d in [r[0] for r in DIVISIONS])
    act_eb_total = sum(fy25_eb(d, 'act_eb') for d in [r[0] for r in DIVISIONS])

    for col, val, fmt in [(3, bud_rev_total, NUM_FMT), (4, act_rev_total, NUM_FMT),
                          (5, act_rev_total - bud_rev_total, NUM_FMT),
                          (6, (act_rev_total - bud_rev_total) / bud_rev_total, PCT_FMT),
                          (7, round(bud_eb_total, 1), NUM_FMT), (8, round(act_eb_total, 1), NUM_FMT),
                          (9, round(act_eb_total - bud_eb_total, 1), NUM_FMT),
                          (10, (act_eb_total - bud_eb_total) / bud_eb_total, PCT_FMT)]:
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = fmt
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
    sc = ws.cell(row=row, column=11, value='RED')
    status_pill(sc, 'red')
    ws.cell(row=row, column=12, value='Group EBITDA miss 17.8% — recovery anchored on 3 Red divisions').font = TOTAL_FONT
    ws.cell(row=row, column=12).fill = TOTAL_FILL

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions['K'].width = 9
    ws.column_dimensions['L'].width = 50
    ws.freeze_panes = 'C5'


# ── Sheet 2: Driver Tree ───────────────────────────────────────────────────
def s2_drivers(ws):
    ws.title = 'Driver Tree'
    ws['A1'] = 'FY2025 EBITDA Variance Driver Tree — 5 Worst Divisions'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:G1')
    ws['A2'] = 'Each driver category broken into 4 sub-drivers with EBITDA impact (MYR M)'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:G2')

    headers = ['Division', 'Driver Category', 'Sub-driver', 'EBITDA Impact (MYR M)', 'Volume / Price / Cost / FX', 'Status', 'Action Owner']
    for i, h in enumerate(headers, 1):
        style_h(ws.cell(row=4, column=i, value=h))

    DRIVERS = {
        'prop': [
            ('Project Delivery', [
                ('Mid-Valley Phase 2 completion delay (8 mo)', -180, 'Volume', 'red', 'Div CEO + Project Director'),
                ('Penang Tower scope creep + writedown', -210, 'Cost', 'red', 'Div CFO + GC'),
                ('Iskandar mall fit-out cost overrun', -45, 'Cost', 'amber', 'Construction Head'),
                ('JB town-house phase 1 delay (3 mo)', -28, 'Volume', 'amber', 'Project Manager'),
            ]),
            ('Sales Velocity', [
                ('Klang Valley unsold inventory (320 units)', -95, 'Volume', 'red', 'Sales Director'),
                ('Pricing erosion (5-7% to clear)', -68, 'Price', 'amber', 'Sales Director'),
                ('Loan rejection rate up to 24% (vs 16% norm)', -42, 'Volume', 'amber', 'Sales Director + FS partner'),
                ('Foreign-buyer demand softness', -22, 'Volume', 'amber', 'International Sales'),
            ]),
            ('Asset Management', [
                ('Holding cost on unsold inventory', -38, 'Cost', 'red', 'Asset Director'),
                ('Property tax escalation (KL revaluation)', -12, 'Cost', 'amber', 'Tax + Asset'),
                ('Maintenance & service-charge subsidy', -15, 'Cost', 'amber', 'Facilities'),
                ('Insurance premium hardening', -6, 'Cost', 'green', 'Risk + Insurance'),
            ]),
        ],
        'hc': [
            ('Volume', [
                ('Outpatient volume −9% YoY', -38, 'Volume', 'red', 'Hospital CEOs'),
                ('Inpatient days −11% YoY', -42, 'Volume', 'red', 'Hospital CEOs'),
                ('Surgical case volume −7% YoY', -22, 'Volume', 'amber', 'Specialist Heads'),
                ('Diagnostic test volume +3%', +6, 'Volume', 'green', 'Diagnostic Director'),
            ]),
            ('Price/Mix', [
                ('Insurance reimbursement pressure', -18, 'Price', 'amber', 'Payor Relations'),
                ('Specialist consultation fee compression', -14, 'Price', 'amber', 'Pricing'),
                ('Pharmacy margin compression', -11, 'Mix', 'amber', 'Pharmacy'),
                ('Premium-pay specialist contracts', -28, 'Cost', 'red', 'CHRO + Hospital CEOs'),
            ]),
            ('Cost', [
                ('Subang facility under-utilisation (58%)', -32, 'Cost', 'red', 'Hospital CEO Subang'),
                ('Equipment lease cost step-up', -16, 'Cost', 'amber', 'Procurement'),
                ('Bad debt provisioning increase', -14, 'Cost', 'amber', 'Finance'),
                ('Insurance benefit cost (workforce)', -8, 'Cost', 'amber', 'CHRO'),
            ]),
        ],
        'retail': [
            ('Footfall', [
                ('Flagship store footfall −14%', -52, 'Volume', 'red', 'Retail Operations'),
                ('Online cannibalisation of physical', -36, 'Volume', 'amber', 'Digital Director'),
                ('Tourist segment recovery slower', -22, 'Volume', 'amber', 'Marketing'),
                ('Mid-tier mall vacancy adjacency', -14, 'Volume', 'amber', 'Real Estate'),
            ]),
            ('Inventory', [
                ('Inventory writedown (FW2024 carryover)', -110, 'Cost', 'red', 'Buying + Finance'),
                ('Markdown depth on slow movers', -28, 'Price', 'amber', 'Buying'),
                ('Aged stock shrinkage (quarterly)', -6, 'Cost', 'amber', 'Loss Prevention'),
                ('Stock-mix imbalance (Cat A surplus)', -18, 'Mix', 'amber', 'Buying'),
            ]),
            ('Cost Base', [
                ('Lease cost inflation (CPI+linked)', -22, 'Cost', 'amber', 'Real Estate'),
                ('Marketing campaign over-spend', -18, 'Cost', 'amber', 'CMO'),
                ('Buying margin compression (vendor ask)', -16, 'Mix', 'amber', 'Buying'),
                ('Loyalty programme cost (refresh)', -8, 'Cost', 'green', 'CRM Director'),
            ]),
        ],
        'chem': [
            ('Demand', [
                ('China olefins demand softness', -75, 'Volume', 'red', 'Sales Director'),
                ('Specialty chemicals customer cancellations', -32, 'Volume', 'amber', 'Specialty Sales'),
                ('Aerospace contract shipment delay', -8, 'Volume', 'amber', 'Strategic Accounts'),
                ('Domestic demand resilient', +12, 'Volume', 'green', 'Domestic Sales'),
            ]),
            ('Cost', [
                ('Petrochemical feedstock spike (Q3)', -45, 'Cost', 'amber', 'Procurement'),
                ('Energy & utilities cost step-up', -18, 'Cost', 'amber', 'Operations'),
                ('Plant utilisation drop at Penang', -12, 'Cost', 'amber', 'Plant Manager Penang'),
                ('Logistics & freight cost', -8, 'Cost', 'green', 'Logistics'),
            ]),
            ('FX & Mix', [
                ('MYR/USD adverse FX (export book)', -15, 'FX', 'amber', 'Treasury'),
                ('Sales mix shift to commodity grade', -22, 'Mix', 'amber', 'Sales Director'),
                ('Hedging gain (USD-denom revenue)', +8, 'FX', 'green', 'Treasury'),
                ('Specialty pricing discipline', +6, 'Price', 'green', 'Specialty Sales'),
            ]),
        ],
        'agri': [
            ('Commodity Price', [
                ('CPO price decline 18% YoY', -120, 'Price', 'red', 'Commercial Director'),
                ('Soft fundamentals (Indonesian supply)', -28, 'Price', 'amber', 'Strategy'),
                ('Refinery margin compression', -12, 'Mix', 'amber', 'Refinery Manager'),
                ('PK and PFAD stable', +4, 'Price', 'green', 'Commercial'),
            ]),
            ('Yield', [
                ('Yield decline at Sumatra estates', -34, 'Volume', 'red', 'Estate Heads'),
                ('Replanting capex acceleration impact', -18, 'Volume', 'amber', 'Plantation Director'),
                ('Labour cost inflation (Indonesia)', -22, 'Cost', 'amber', 'CHRO + Estate'),
                ('Disease management programme cost', -8, 'Cost', 'amber', 'Agronomy'),
            ]),
            ('Regulatory & FX', [
                ('Indonesian export levy increase', -38, 'Cost', 'red', 'Govt Relations'),
                ('IDR weakening vs USD', -16, 'FX', 'amber', 'Treasury'),
                ('NDPE compliance cost', -6, 'Cost', 'green', 'Sustainability'),
                ('Carbon-credit revenue (initial)', +12, 'Price', 'green', 'Sustainability'),
            ]),
        ],
    }

    row = 5
    for div_id, drivers in DRIVERS.items():
        div_name = next(d[1] for d in DIVISIONS if d[0] == div_id)
        # Division header band
        c = ws.cell(row=row, column=1, value=div_name)
        c.font = SUB_FONT
        c.fill = SUB_FILL
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = SUB_FILL
        row += 1
        for cat_name, sub_drivers in drivers:
            for sd_name, impact, attr, status, owner in sub_drivers:
                ws.cell(row=row, column=1, value='').font = BODY_FONT
                ws.cell(row=row, column=2, value=cat_name).font = BODY_FONT
                ws.cell(row=row, column=3, value=sd_name).font = BODY_FONT
                c = ws.cell(row=row, column=4, value=impact)
                c.number_format = NUM_FMT
                if impact < 0:
                    c.font = Font(color='C00000', size=10, name='Aptos', bold=abs(impact) > 50)
                else:
                    c.font = Font(color='00B050', size=10, name='Aptos')
                ws.cell(row=row, column=5, value=attr).font = BODY_FONT
                sc = ws.cell(row=row, column=6, value=status.upper())
                status_pill(sc, status)
                sc.alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=7, value=owner).font = BODY_FONT
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = BORDER
                row += 1
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 30
    ws.freeze_panes = 'B5'


# ── Sheet 3: Monthly Variance Trend ────────────────────────────────────────
def s3_monthly(ws):
    ws.title = 'Monthly Variance Trend'
    ws['A1'] = 'FY2025 Monthly Variance Trend — Budget vs Actual EBITDA (MYR M)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:N1')
    ws['A2'] = '11 divisions × 12 months × 4 metrics (Bud, Act, Var, Var%)'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:N2')

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    headers = ['Division', 'Metric'] + months
    for i, h in enumerate(headers, 1):
        style_h(ws.cell(row=4, column=i, value=h))

    METRICS = [
        ('Budget EBITDA (MYR M)', 'bud'),
        ('Actual EBITDA (MYR M)', 'act'),
        ('Variance (MYR M)', 'var'),
        ('Variance %', 'pct'),
    ]

    # Seasonality factors per division (sums to 1.0 across 12 months)
    SEASONALITY = {
        'chem':   [0.07,0.07,0.08,0.08,0.09,0.09,0.09,0.09,0.09,0.09,0.08,0.08],
        'agri':   [0.07,0.07,0.08,0.08,0.08,0.08,0.08,0.09,0.09,0.10,0.10,0.08],
        'mfg':    [0.08,0.07,0.08,0.08,0.09,0.08,0.09,0.09,0.09,0.09,0.08,0.08],
        'fs':     [0.08,0.08,0.09,0.08,0.08,0.08,0.08,0.08,0.08,0.09,0.09,0.09],
        'prop':   [0.06,0.07,0.08,0.07,0.07,0.07,0.08,0.09,0.09,0.10,0.11,0.11],  # Q4 worst; concentrated writedowns
        'bpo':    [0.08,0.08,0.08,0.08,0.08,0.08,0.09,0.09,0.09,0.09,0.08,0.08],
        'trade':  [0.08,0.08,0.08,0.08,0.08,0.08,0.08,0.10,0.12,0.10,0.06,0.06],  # Hindustan Bulk in Sep
        'pharma': [0.08,0.08,0.08,0.08,0.08,0.09,0.09,0.09,0.09,0.09,0.08,0.07],
        'hc':     [0.07,0.07,0.08,0.08,0.08,0.08,0.09,0.09,0.09,0.09,0.09,0.09],
        'retail': [0.06,0.06,0.07,0.07,0.07,0.07,0.08,0.09,0.09,0.10,0.12,0.12],  # Q4 worst
        'treas':  [0.08,0.08,0.08,0.08,0.08,0.08,0.08,0.09,0.09,0.10,0.09,0.07],
    }
    # Variance pattern — flat improvement vs deterioration
    VAR_PATTERN = {
        'chem':   [0.95,0.92,0.93,0.91,0.88,0.85,0.85,0.84,0.86,0.88,0.92,0.96],
        'agri':   [0.96,0.94,0.92,0.90,0.88,0.86,0.86,0.84,0.86,0.88,0.92,0.94],
        'mfg':    [0.97,0.95,0.93,0.92,0.91,0.89,0.92,0.94,0.96,0.97,0.98,0.99],
        'fs':     [1.05,1.06,1.07,1.07,1.07,1.07,1.08,1.08,1.08,1.07,1.07,1.07],
        'prop':   [0.50,0.40,0.20,0.10,-0.10,-0.20,-0.40,-0.50,-0.70,-0.85,-1.00,-1.10],
        'bpo':    [1.02,1.03,1.04,1.05,1.04,1.04,1.04,1.05,1.05,1.04,1.04,1.04],
        'trade':  [0.98,0.95,0.92,0.90,0.88,0.85,0.82,0.65,0.40,0.85,0.95,0.98],
        'pharma': [1.00,0.98,0.98,0.97,0.97,0.96,0.96,0.95,0.95,0.96,0.97,0.98],
        'hc':     [0.85,0.75,0.65,0.50,0.30,0.10,-0.10,-0.30,-0.45,-0.55,-0.65,-0.70],
        'retail': [0.80,0.70,0.55,0.40,0.20,0.00,-0.20,-0.40,-0.55,-0.70,-0.90,-1.00],
        'treas':  [0.90,0.88,0.86,0.84,0.82,0.80,0.78,0.76,0.74,0.72,0.70,0.68],
    }

    row = 5
    for did, dname, _, _, _ in DIVISIONS:
        bud_total = fy25_eb(did, 'bud_eb')
        act_total = fy25_eb(did, 'act_eb')
        for mi, (mlabel, mkey) in enumerate(METRICS):
            ws.cell(row=row, column=1, value=dname if mi == 0 else '').font = Font(bold=mi==0, size=10, name='Aptos')
            ws.cell(row=row, column=2, value=mlabel).font = BODY_FONT
            for ci, m in enumerate(months):
                bud_m = bud_total * SEASONALITY[did][ci]
                act_m = bud_m * VAR_PATTERN[did][ci]
                var_m = act_m - bud_m
                pct_m = var_m / bud_m if bud_m else 0
                vals = {'bud': bud_m, 'act': act_m, 'var': var_m, 'pct': pct_m}
                cell = ws.cell(row=row, column=3 + ci, value=round(vals[mkey], 1) if mkey != 'pct' else round(vals[mkey], 4))
                cell.number_format = PCT_FMT if mkey == 'pct' else NUM_FMT
                cell.font = BODY_FONT
                if mkey in ('var','pct') and vals[mkey] < 0:
                    cell.font = Font(color='C00000', size=10, name='Aptos', bold=abs(vals[mkey]) > (10 if mkey == 'var' else 0.1))
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 24
    for col in range(3, 15):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.freeze_panes = 'C5'


# ── Sheet 4: Recovery Levers ───────────────────────────────────────────────
def s4_levers(ws):
    ws.title = 'Recovery Levers'
    ws['A1'] = 'FY2026 Recovery Levers — Specific Actions to Recover the EBITDA Gap'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:H1')
    ws['A2'] = '32 actions × 6 attributes (owner, expected impact, time-to-impact, confidence, dependencies, status)'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:H2')

    headers = ['#', 'Division', 'Lever', 'Owner', 'FY26 Impact (MYR M)', 'Time-to-impact', 'Confidence', 'Dependencies / Notes']
    for i, h in enumerate(headers, 1):
        style_h(ws.cell(row=4, column=i, value=h))

    LEVERS = [
        # Properties
        ('Properties', 'Mid-Valley Phase 2 sell-down with bundled financing', 'Div CEO Properties', 95, '6-9 mo', 'Medium', 'Co-marketing with Group FS arm; price discipline'),
        ('Properties', 'Penang Tower repositioning to mixed-use', 'Asset Director', 75, '12-18 mo', 'Medium', 'Concept design + planning approval'),
        ('Properties', 'Klang Valley land-bank disposal (3 parcels)', 'Investments Director', 110, '6-12 mo', 'High', 'IM out; auction underway'),
        ('Properties', 'Cost rationalisation programme (sales & marketing)', 'Div CFO Properties', 32, '3-6 mo', 'High', 'Identified savings vs FY25 baseline'),
        ('Properties', 'JB town-house phase 2 launch', 'Sales Director', 28, '9-12 mo', 'Medium', 'Approval secured; market timing critical'),
        # Healthcare
        ('Healthcare', 'Subang outpatient/day-surgery conversion', 'Div CEO Healthcare', 65, '9-12 mo', 'Medium', 'MOH approval + capex MYR 35M'),
        ('Healthcare', 'Specialist mix optimisation (focus high-volume procedures)', 'Specialist Heads', 38, '6-9 mo', 'Medium', 'Specialist contract renegotiations'),
        ('Healthcare', 'Insurance bundled product (with FS arm)', 'CMO + FS partner', 22, '6-12 mo', 'Low', 'Cross-divisional governance'),
        ('Healthcare', 'Bad-debt programme tightening', 'Finance + Billing', 14, '3-6 mo', 'High', 'Collections process redesign'),
        ('Healthcare', 'Pharmacy retail expansion within network', 'Pharmacy Director', 18, '6-9 mo', 'High', 'Retail format established'),
        # Retail
        ('Retail', 'Close 18 underperforming stores', 'Real Estate Director', 48, '6-12 mo', 'High', 'Lease exit negotiations underway'),
        ('Retail', 'Omni-channel platform deployment', 'Digital Director', 36, '6-12 mo', 'Medium', 'IT investment MYR 28M'),
        ('Retail', 'Beauty + electronics category-killer pilot (3 stores)', 'Format Innovation', 28, '9-12 mo', 'Medium', 'Buying realignment + format design'),
        ('Retail', 'Dynamic pricing platform', 'Pricing Director', 22, '6-9 mo', 'Medium', 'Retailux platform deploy'),
        ('Retail', 'Loyalty programme refresh + fintech instalment', 'CRM Director', 18, '6-9 mo', 'Medium', 'FS partner integration'),
        # Chemicals
        ('Chemicals', 'Specialty chemicals capacity expansion at Pengerang', 'Plant Manager Pengerang', 55, '12-18 mo', 'High', 'Capex MYR 240M; brownfield path'),
        ('Chemicals', 'Methane abatement + 100MW captive solar', 'Plant Manager All', 22, '12-18 mo', 'High', 'Sustainability-linked sukuk drawdown'),
        ('Chemicals', 'EV battery materials qualification (3 OEM)', 'Specialty Sales', 38, '12-15 mo', 'Medium', 'Qualification cycles + samples'),
        # Agribusiness
        ('Agribusiness', 'Yield recovery programme Sumatra', 'Estate Heads', 32, '9-12 mo', 'Medium', 'Agronomy + soil-health intervention'),
        ('Agribusiness', 'Replanting acceleration funded by green sukuk', 'Plantation Director', 18, '12-24 mo', 'High', 'Funded; execution discipline key'),
        ('Agribusiness', '2 Kalimantan estate disposals + reinvestment', 'Investments + Operations', 28, '12-18 mo', 'High', 'Beauty parade; recycle to refinery'),
        # Manufacturing
        ('Manufacturing', 'Penang plant downtime root-cause programme', 'Plant Manager Penang', 24, '6-9 mo', 'High', 'Mech reliability + spares'),
        ('Manufacturing', 'KL + Subang plant closure → Senai consolidation', 'Operations Director', 38, '12-18 mo', 'High', 'Workforce transition plan'),
        ('Manufacturing', 'AI-assisted predictive maintenance roll-out', 'Digital + Operations', 18, '6-12 mo', 'Medium', 'Group Digital platform partner'),
        # FS - upside levers (compound)
        ('Financial Services', 'Wealth management arm launch', 'Div CEO FS', 90, '12-18 mo', 'High', 'OJK + BNM licence; build advisor team'),
        ('Financial Services', 'Margin lending growth + securities financing', 'Wealth + Capital Markets', 65, '6-12 mo', 'High', 'Already on growth trajectory'),
        # BPO - upside
        ('BPO', 'Two US bolt-on acquisitions', 'Div CEO BPO', 120, '12-24 mo', 'Medium', 'M-1 in DD; M-2 long-list'),
        ('BPO', 'Outcome-based pricing pivot', 'Pricing + Sales', 55, '9-12 mo', 'Medium', 'Top-15 client conversations'),
        # Pharma - upside
        ('Pharmaceuticals', '12 ANDA dossier filings + 5 launches', 'Regulatory + R&D', 65, '12-18 mo', 'Medium', 'BPOM + Pharmacy Board cycles'),
        ('Pharmaceuticals', 'Branded portfolio build (cardio + women\'s health)', 'Marketing + R&D', 40, '12-18 mo', 'Medium', 'One bolt-on (M-5) considered'),
        # Trading
        ('Trading', 'Risk-managed run-off post-Hindustan Bulk', 'Div CEO Trading', 18, '6-12 mo', 'High', 'Counterparty limits + insurance'),
        ('Trading', 'Specialty trading pivot (rare earths, agri-soft)', 'Commercial', 12, '12-18 mo', 'Low', 'Talent recruitment'),
    ]

    row = 5
    for i, (div, lever, owner, impact, ttim, conf, deps) in enumerate(LEVERS, 1):
        ws.cell(row=row, column=1, value=i).font = BODY_FONT
        ws.cell(row=row, column=2, value=div).font = BODY_FONT
        ws.cell(row=row, column=3, value=lever).font = BODY_FONT
        ws.cell(row=row, column=4, value=owner).font = BODY_FONT
        c = ws.cell(row=row, column=5, value=impact)
        c.number_format = NUM_FMT
        c.font = Font(color='00B050', size=10, name='Aptos', bold=impact > 50)
        ws.cell(row=row, column=6, value=ttim).font = BODY_FONT
        sc = ws.cell(row=row, column=7, value=conf)
        if conf == 'High':
            sc.fill = GREEN_FILL
            sc.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
        elif conf == 'Medium':
            sc.fill = AMBER_FILL
            sc.font = Font(bold=True, color='000000', size=10, name='Aptos')
        else:
            sc.fill = RED_FILL
            sc.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
        sc.alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=8, value=deps).font = BODY_FONT
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    # Total row
    ws.cell(row=row, column=1, value='').fill = TOTAL_FILL
    ws.cell(row=row, column=2, value='ALL').font = TOTAL_FONT
    ws.cell(row=row, column=2).fill = TOTAL_FILL
    ws.cell(row=row, column=3, value='Total programme impact').font = TOTAL_FONT
    ws.cell(row=row, column=3).fill = TOTAL_FILL
    ws.cell(row=row, column=4, value='').fill = TOTAL_FILL
    total_impact = sum(l[3] for l in LEVERS)
    c = ws.cell(row=row, column=5, value=total_impact)
    c.number_format = NUM_FMT
    c.font = TOTAL_FONT
    c.fill = TOTAL_FILL
    for col in range(6, 9):
        ws.cell(row=row, column=col).fill = TOTAL_FILL

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 40
    ws.freeze_panes = 'B5'


def main(out_path):
    wb = Workbook()
    wb.remove(wb.active)
    for builder in (s1_summary, s2_drivers, s3_monthly, s4_levers):
        ws = wb.create_sheet()
        builder(ws)
    wb.save(out_path)
    print(f'Wrote {out_path}')


if __name__ == '__main__':
    out = sys.argv[1] if len(sys.argv) > 1 else 'files/04_Zava_Divisional_Variance_FY2025.xlsx'
    main(out)
