"""Generates 01_Zava_Group_Financial_Performance.xlsx — 6 sheets, 4000+ cells.

Sheets:
  1. Group P&L Summary       — 11 div × 5yr × 13 P&L lines
  2. Division Revenue Bridge — 11 div × 4 quarters × 6 bridge components
  3. Quarterly Earnings Tracker — 11 div × 16 quarters × 6 metrics
  4. Balance Sheet Summary   — 11 div × 5 years × 11 BS lines
  5. Key Financial Ratios    — group + 11 div × 5 years × 12 ratios
  6. Variance Drill-Down     — 11 div × top 8 variance drivers × 5 impacts
"""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from gen_ref_files.common import DIVISIONS, DIVISION_IDS, FY25, fy25_eb, hist_rev, hist_eb_pct, hist_eb, status_color, YEARS, HISTORY

# ── styling helpers ────────────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Aptos')
SUBHEADER_FILL = PatternFill('solid', fgColor='2E75B6')
SUBHEADER_FONT = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
TOTAL_FILL = PatternFill('solid', fgColor='D9E1F2')
TOTAL_FONT = Font(bold=True, size=10, name='Aptos')
BODY_FONT = Font(size=10, name='Aptos')
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT = '#,##0;[Red](#,##0);-'
PCT_FMT = '0.0%;[Red]-0.0%;-'
RATIO_FMT = '0.00'


def style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER


def write_band(ws, row, cols, fill, font=None):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        if font:
            cell.font = font


# ── Sheet 1: Group P&L Summary ─────────────────────────────────────────────
def build_pl_summary(ws):
    ws.title = 'Group P&L Summary'
    ws['A1'] = 'Zava Group — Consolidated P&L Summary (MYR Million)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:G1')
    ws['A2'] = 'FY2022A — FY2026F | Source: Group FP&A | Status: BOARD-CONFIDENTIAL'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:G2')

    # Header row
    headers = ['Division'] + YEARS + ['CAGR FY22-26']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_header(c)

    PL_LINES = [
        ('Revenue', 'rev', NUM_FMT),
        ('  Cost of Goods Sold', 'cogs', NUM_FMT),
        ('Gross Profit', 'gp', NUM_FMT),
        ('  Gross Margin %', 'gm_pct', PCT_FMT),
        ('Operating Expenses', 'opex', NUM_FMT),
        ('EBITDA', 'ebitda', NUM_FMT),
        ('  EBITDA Margin %', 'eb_pct', PCT_FMT),
        ('Depreciation & Amortisation', 'da', NUM_FMT),
        ('EBIT', 'ebit', NUM_FMT),
        ('Finance Costs (net)', 'fin', NUM_FMT),
        ('PBT', 'pbt', NUM_FMT),
        ('Tax', 'tax', NUM_FMT),
        ('PAT', 'pat', NUM_FMT),
        ('Minority Interest', 'mi', NUM_FMT),
        ('PATAMI', 'patami', NUM_FMT),
    ]

    row = 6
    group_totals = {y: {} for y in YEARS}
    for did, dname, _, _, _ in DIVISIONS:
        # Division header band
        c = ws.cell(row=row, column=1, value=dname)
        c.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
        c.fill = SUBHEADER_FILL
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = SUBHEADER_FILL
        row += 1

        # Compute P&L for each year for this division
        for line_label, line_key, fmt in PL_LINES:
            ws.cell(row=row, column=1, value=line_label).font = BODY_FONT
            for yi, year in enumerate(YEARS):
                rev = HISTORY[did][0][yi]
                eb_pct = HISTORY[did][1][yi]
                ebitda = rev * eb_pct
                # Construct full P&L from rev + ebitda
                cogs_pct = 0.62 + 0.02 * (1 if did in ('trade','retail') else 0) - 0.05 * (1 if did in ('fs','bpo','treas') else 0)
                cogs = -rev * cogs_pct
                gp = rev + cogs
                gm_pct = gp / rev if rev else 0
                opex = ebitda - gp  # opex (negative)
                da = -rev * 0.045 if did != 'treas' else -rev * 0.02
                ebit = ebitda + da
                fin = -rev * 0.025 if did != 'fs' else rev * 0.04
                pbt = ebit + fin
                tax = -pbt * 0.24 if pbt > 0 else 0
                pat = pbt + tax
                mi = -pat * 0.08 if pat > 0 else 0
                patami = pat + mi

                vals = {
                    'rev': rev, 'cogs': cogs, 'gp': gp, 'gm_pct': gm_pct,
                    'opex': opex, 'ebitda': ebitda, 'eb_pct': eb_pct, 'da': da,
                    'ebit': ebit, 'fin': fin, 'pbt': pbt, 'tax': tax,
                    'pat': pat, 'mi': mi, 'patami': patami,
                }
                v = vals[line_key]
                cell = ws.cell(row=row, column=2 + yi, value=round(v, 1))
                cell.number_format = fmt
                cell.font = BODY_FONT

                # Aggregate for group totals
                group_totals[year].setdefault(line_key, 0)
                group_totals[year][line_key] += v

            # CAGR (revenue line and EBITDA line only)
            if line_key in ('rev', 'ebitda'):
                v22 = HISTORY[did][0][0] * (1 if line_key == 'rev' else HISTORY[did][1][0])
                v26 = HISTORY[did][0][4] * (1 if line_key == 'rev' else HISTORY[did][1][4])
                if v22 > 0 and v26 > 0:
                    cagr = (v26 / v22) ** 0.25 - 1
                    cell = ws.cell(row=row, column=7, value=cagr)
                    cell.number_format = PCT_FMT
                    cell.font = BODY_FONT

            row += 1
        row += 1  # gap row

    # Group total band
    c = ws.cell(row=row, column=1, value='ZAVA GROUP CONSOLIDATED')
    c.font = Font(bold=True, color='FFFFFF', size=11, name='Aptos')
    c.fill = HEADER_FILL
    for col in range(2, 8):
        ws.cell(row=row, column=col).fill = HEADER_FILL
    row += 1
    for line_label, line_key, fmt in PL_LINES:
        ws.cell(row=row, column=1, value=line_label).font = TOTAL_FONT
        ws.cell(row=row, column=1).fill = TOTAL_FILL
        for yi, year in enumerate(YEARS):
            v = group_totals[year].get(line_key, 0)
            if line_key == 'gm_pct':
                # recompute group-level percentage from rev+gp
                rev_g = group_totals[year]['rev']
                gp_g = group_totals[year]['gp']
                v = gp_g / rev_g if rev_g else 0
            elif line_key == 'eb_pct':
                rev_g = group_totals[year]['rev']
                eb_g = group_totals[year]['ebitda']
                v = eb_g / rev_g if rev_g else 0
            cell = ws.cell(row=row, column=2 + yi, value=round(v, 1) if line_key not in ('gm_pct','eb_pct') else round(v, 4))
            cell.number_format = fmt
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 32
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = 'B5'


# ── Sheet 2: Division Revenue Bridge ───────────────────────────────────────
def build_revenue_bridge(ws):
    ws.title = 'Division Revenue Bridge'
    ws['A1'] = 'Division Revenue Bridge — FY2024A → FY2025A (MYR Million)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:H1')
    ws['A2'] = 'Bridge components: Organic, M&A, FX, Disposals, Pricing, Volume | Source: Group FP&A'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:H2')

    headers = ['Division', 'FY2024A Rev', 'Organic', 'M&A', 'FX', 'Disposals', 'Pricing', 'Volume', 'FY2025A Rev', 'Total Bridge', 'Variance vs Bud']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_header(c)

    # Bridge data — engineered so FY24 + bridge ≈ FY25
    BRIDGE = {
        'chem':   ( 100,  150,  -180,    0,  -50,   30),  # Organic, M&A, FX, Disposals, Pricing, Volume
        'agri':   (  80,    0,  -210,    0, -120,   50),
        'mfg':    ( 130,   90,   -50,    0,  -40,   70),
        'fs':     ( 320,   60,    20,    0,    0,    0),
        'prop':   (-280,    0,    0,  -200, -100,  -20),
        'bpo':    ( 220,   80,   30,    0,    0,   20),
        'trade':  (-150,    0, -210,    0,  -40,    0),
        'pharma': (  40,    0,    0,    0,   10,    0),
        'hc':     (-120,    0,    0,  -50,  -20,  -10),
        'retail': (-150,    0,    0,    0,  -30,  -20),
        'treas':  (  -5,    0,    0,    0,    0,    0),
    }
    row = 6
    for did, dname, _, _, _ in DIVISIONS:
        ws.cell(row=row, column=1, value=dname).font = BODY_FONT
        fy24 = HISTORY[did][0][2]
        fy25 = HISTORY[did][0][3]
        bridge = BRIDGE[did]
        ws.cell(row=row, column=2, value=fy24).number_format = NUM_FMT
        for i, v in enumerate(bridge):
            cell = ws.cell(row=row, column=3 + i, value=v)
            cell.number_format = NUM_FMT
            if v < 0:
                cell.font = Font(color='C00000', size=10, name='Aptos')
            else:
                cell.font = BODY_FONT
        ws.cell(row=row, column=9, value=fy25).number_format = NUM_FMT
        ws.cell(row=row, column=10, value=sum(bridge)).number_format = NUM_FMT
        bud25 = FY25[did][0]
        var = fy25 - bud25
        c = ws.cell(row=row, column=11, value=var)
        c.number_format = NUM_FMT
        c.font = Font(color='C00000' if var < 0 else '00B050', size=10, name='Aptos', bold=abs(var) > 200)
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    # Group total
    ws.cell(row=row, column=1, value='GROUP TOTAL').font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    for col_idx, col_name in enumerate(['fy24','organic','ma','fx','disp','price','vol','fy25','total','var'], 2):
        if col_name == 'fy24':
            v = sum(HISTORY[d][0][2] for d in DIVISION_IDS)
        elif col_name == 'fy25':
            v = sum(HISTORY[d][0][3] for d in DIVISION_IDS)
        elif col_name == 'total':
            v = sum(sum(BRIDGE[d]) for d in DIVISION_IDS)
        elif col_name == 'var':
            v = sum(HISTORY[d][0][3] - FY25[d][0] for d in DIVISION_IDS)
        else:
            idx = ['organic','ma','fx','disp','price','vol'].index(col_name)
            v = sum(BRIDGE[d][idx] for d in DIVISION_IDS)
        c = ws.cell(row=row, column=col_idx, value=v)
        c.number_format = NUM_FMT
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL

    ws.column_dimensions['A'].width = 26
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.freeze_panes = 'B5'


# ── Sheet 3: Quarterly Earnings Tracker ────────────────────────────────────
def build_quarterly_tracker(ws):
    ws.title = 'Quarterly Earnings Tracker'
    ws['A1'] = 'Quarterly Earnings Tracker — Q1 FY2022 to Q4 FY2025 (MYR Million)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:R1')
    ws['A2'] = '11 divisions × 16 quarters × 6 metrics | Source: Group FP&A'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:R2')

    # Quarter headers
    quarters = [f'Q{q} FY{y}' for y in (2022, 2023, 2024, 2025) for q in (1, 2, 3, 4)]
    headers = ['Division', 'Metric'] + quarters
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_header(c)

    METRICS = [
        ('Revenue (MYR M)', 'rev'),
        ('EBITDA (MYR M)', 'eb'),
        ('EBITDA Margin %', 'eb_pct'),
        ('Free Cash Flow (MYR M)', 'fcf'),
        ('Capex (MYR M)', 'capex'),
        ('Headcount (FTE)', 'hc'),
    ]

    HC_BASE = {'chem':2400,'agri':3100,'mfg':2800,'fs':900,'prop':650,'bpo':5200,'trade':380,'pharma':1700,'hc':1900,'retail':2400,'treas':75}

    row = 6
    for did, dname, _, _, _ in DIVISIONS:
        # Each division spans 6 metric rows, so band the first row only.
        for mi, (mlabel, mkey) in enumerate(METRICS):
            ws.cell(row=row, column=1, value=dname if mi == 0 else '').font = Font(bold=mi==0, size=10, name='Aptos')
            ws.cell(row=row, column=2, value=mlabel).font = BODY_FONT
            # 16 quarters of values (interpolate from annual figures)
            for qi, qlabel in enumerate(quarters):
                yi = qi // 4  # 0..3 (FY22..FY25)
                qpos = qi % 4  # 0..3 (Q1..Q4)
                # Quarterly is annual / 4 with seasonality / drift
                annual_rev = HISTORY[did][0][yi]
                annual_eb = HISTORY[did][0][yi] * HISTORY[did][1][yi]
                # Seasonality: Q1 87%, Q2 95%, Q3 105%, Q4 113% (sum=400%)
                seasonality = [0.87, 0.95, 1.05, 1.13][qpos]
                # Q4 FY25 hit hardest for Red divisions
                if did in ('prop','hc','retail') and yi == 3 and qpos == 3:
                    seasonality *= 0.55  # accelerated deterioration in Q4
                rev_q = annual_rev / 4 * seasonality
                eb_q = annual_eb / 4 * seasonality * (0.8 if (did in ('prop','hc','retail') and yi == 3 and qpos == 3) else 1.0)
                eb_pct_q = eb_q / rev_q if rev_q else 0
                fcf_q = eb_q * 0.55 - rev_q * 0.04
                capex_q = annual_rev / 4 * 0.05
                hc_q = HC_BASE[did] * (1 + 0.02 * yi) * (1.02 if did in ('bpo','fs') else 1.0)

                vals = {'rev': round(rev_q,1), 'eb': round(eb_q,1), 'eb_pct': round(eb_pct_q,4),
                        'fcf': round(fcf_q,1), 'capex': round(-abs(capex_q),1), 'hc': int(hc_q)}
                cell = ws.cell(row=row, column=3 + qi, value=vals[mkey])
                cell.number_format = PCT_FMT if mkey == 'eb_pct' else NUM_FMT
                cell.font = BODY_FONT
                if mkey in ('eb','eb_pct') and vals[mkey] < 0:
                    cell.font = Font(color='C00000', size=10, name='Aptos', bold=True)
            row += 1
        row += 1  # gap row between divisions

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    for col in range(3, 19):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.freeze_panes = 'C5'


# ── Sheet 4: Balance Sheet Summary ─────────────────────────────────────────
def build_balance_sheet(ws):
    ws.title = 'Balance Sheet Summary'
    ws['A1'] = 'Balance Sheet Summary by Division — FY2022A to FY2026F (MYR Million)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:G1')
    ws['A2'] = 'Source: Group FP&A | Year-end balances'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:G2')

    headers = ['Division / Line Item'] + YEARS + ['FY26F vs FY24A']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_header(c)

    BS_LINES = [
        ('Total Assets', 'ta'),
        ('  Cash & Equivalents', 'cash'),
        ('  Trade Receivables', 'ar'),
        ('  Inventory', 'inv'),
        ('  Property, Plant & Equipment', 'ppe'),
        ('  Intangibles & Goodwill', 'intang'),
        ('Total Liabilities', 'tl'),
        ('  Long-term Debt', 'lt_debt'),
        ('  Short-term Debt', 'st_debt'),
        ('  Trade Payables', 'ap'),
        ('Total Equity', 'eq'),
    ]

    row = 6
    group_totals = {y: {} for y in YEARS}
    for did, dname, _, _, _ in DIVISIONS:
        c = ws.cell(row=row, column=1, value=dname)
        c.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
        c.fill = SUBHEADER_FILL
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = SUBHEADER_FILL
        row += 1

        # Asset/liability scaling — conglomerate-typical ratios off revenue
        for line_label, line_key in BS_LINES:
            ws.cell(row=row, column=1, value=line_label).font = BODY_FONT
            for yi, year in enumerate(YEARS):
                rev = HISTORY[did][0][yi]
                eb_pct = HISTORY[did][1][yi]
                # Asset intensity varies by sector
                if did in ('prop','mfg','chem','agri'):
                    asset_intensity = 1.20  # high
                elif did in ('hc','retail','pharma'):
                    asset_intensity = 0.85
                elif did in ('fs',):
                    asset_intensity = 4.50  # FS holds large investment book
                elif did == 'bpo':
                    asset_intensity = 0.55
                elif did == 'treas':
                    asset_intensity = 8.00  # holdings company
                else:
                    asset_intensity = 0.65
                ta = rev * asset_intensity
                cash = ta * (0.18 if did == 'fs' else 0.07)
                ar = rev * (0.13 if did == 'trade' else 0.10)
                inv = rev * (0.18 if did in ('chem','agri','mfg','retail') else 0.08)
                ppe = ta * (0.55 if did in ('prop','mfg','hc') else 0.32)
                intang = ta * (0.10 if did in ('pharma','bpo') else 0.04)
                # Liabilities — conglomerate carries ~40-60% leverage
                lev = 0.55 if did != 'treas' else 0.30
                tl = ta * lev
                lt_debt = tl * 0.55
                st_debt = tl * 0.18
                ap = rev * 0.11
                eq = ta - tl

                vals = {'ta':ta,'cash':cash,'ar':ar,'inv':inv,'ppe':ppe,'intang':intang,'tl':tl,'lt_debt':lt_debt,'st_debt':st_debt,'ap':ap,'eq':eq}
                v = vals[line_key]
                cell = ws.cell(row=row, column=2 + yi, value=round(v, 1))
                cell.number_format = NUM_FMT
                cell.font = BODY_FONT
                group_totals[year].setdefault(line_key, 0)
                group_totals[year][line_key] += v

            # FY26F vs FY24A delta
            v24 = group_totals[YEARS[2]].get(line_key, 0) if False else None  # placeholder, recalc below
            # Skip for division rows — only at group level
            row += 1
        row += 1

    # Group consolidation
    c = ws.cell(row=row, column=1, value='ZAVA GROUP CONSOLIDATED')
    c.font = Font(bold=True, color='FFFFFF', size=11, name='Aptos')
    c.fill = HEADER_FILL
    for col in range(2, 8):
        ws.cell(row=row, column=col).fill = HEADER_FILL
    row += 1
    for line_label, line_key in BS_LINES:
        ws.cell(row=row, column=1, value=line_label).font = TOTAL_FONT
        ws.cell(row=row, column=1).fill = TOTAL_FILL
        for yi, year in enumerate(YEARS):
            v = group_totals[year].get(line_key, 0)
            cell = ws.cell(row=row, column=2 + yi, value=round(v, 1))
            cell.number_format = NUM_FMT
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
        # FY26F vs FY24A delta
        delta = group_totals['FY2026F'].get(line_key, 0) - group_totals['FY2024A'].get(line_key, 0)
        c = ws.cell(row=row, column=7, value=round(delta, 1))
        c.number_format = NUM_FMT
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        row += 1

    ws.column_dimensions['A'].width = 30
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = 'B5'


# ── Sheet 5: Key Financial Ratios ─────────────────────────────────────────
def build_ratios(ws):
    ws.title = 'Key Financial Ratios'
    ws['A1'] = 'Key Financial Ratios — Group + 11 Divisions × FY2022A to FY2026F'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:G1')
    ws['A2'] = '12 ratios per entity per year | Source: Group FP&A'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:G2')

    headers = ['Entity', 'Ratio'] + YEARS
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_header(c)

    RATIOS = [
        ('Net Debt / EBITDA', 'nd_eb', RATIO_FMT),
        ('Interest Cover (×)', 'int_cov', RATIO_FMT),
        ('Current Ratio (×)', 'cur', RATIO_FMT),
        ('Quick Ratio (×)', 'quick', RATIO_FMT),
        ('Return on Assets', 'roa', PCT_FMT),
        ('Return on Equity', 'roe', PCT_FMT),
        ('Return on Invested Capital', 'roic', PCT_FMT),
        ('Asset Turnover (×)', 'at', RATIO_FMT),
        ('Gross Margin %', 'gm', PCT_FMT),
        ('EBITDA Margin %', 'eb_pct', PCT_FMT),
        ('Net Margin %', 'nm', PCT_FMT),
        ('Liquidity Coverage (Cash / ST Debt)', 'liq', RATIO_FMT),
    ]

    row = 6
    # First do the Group row (aggregate), then 11 divisions
    for entity_idx, (entity, key) in enumerate([('Zava Group Consolidated', 'GROUP')] + [(DIVISIONS[i][1], DIVISIONS[i][0]) for i in range(11)]):
        for ri, (rlabel, rkey, fmt) in enumerate(RATIOS):
            ws.cell(row=row, column=1, value=entity if ri == 0 else '').font = Font(bold=ri==0, size=10, name='Aptos')
            ws.cell(row=row, column=2, value=rlabel).font = BODY_FONT
            for yi, year in enumerate(YEARS):
                # Compute ratio
                if key == 'GROUP':
                    rev = sum(HISTORY[d][0][yi] for d in DIVISION_IDS)
                    eb = sum(HISTORY[d][0][yi] * HISTORY[d][1][yi] for d in DIVISION_IDS)
                else:
                    rev = HISTORY[key][0][yi]
                    eb = HISTORY[key][0][yi] * HISTORY[key][1][yi]
                # Synthetic balance-sheet derivations consistent with sheet 4
                if key == 'GROUP':
                    ta = sum(HISTORY[d][0][yi] * (1.20 if d in ('prop','mfg','chem','agri') else 0.85 if d in ('hc','retail','pharma') else 4.50 if d=='fs' else 0.55 if d=='bpo' else 8.00 if d=='treas' else 0.65) for d in DIVISION_IDS)
                else:
                    ai = (1.20 if key in ('prop','mfg','chem','agri') else 0.85 if key in ('hc','retail','pharma') else 4.50 if key=='fs' else 0.55 if key=='bpo' else 8.00 if key=='treas' else 0.65)
                    ta = rev * ai
                tl = ta * 0.55
                lt_debt = tl * 0.55
                st_debt = tl * 0.18
                cash = ta * (0.18 if key == 'fs' else 0.07 if key != 'GROUP' else 0.10)
                eq = ta - tl
                net_debt = lt_debt + st_debt - cash
                fin_cost = rev * 0.025
                pat = max(eb * 0.55, 0)  # crude approximation

                v = {
                    'nd_eb': net_debt / eb if eb > 0 else 99.9,
                    'int_cov': eb / fin_cost if fin_cost > 0 else 99.9,
                    'cur': (cash + ta * 0.20) / (st_debt + rev * 0.11) if (st_debt + rev * 0.11) else 0,
                    'quick': cash / (st_debt + rev * 0.11) if (st_debt + rev * 0.11) else 0,
                    'roa': pat / ta if ta else 0,
                    'roe': pat / eq if eq > 0 else 0,
                    'roic': eb / (eq + lt_debt + st_debt) if (eq + lt_debt + st_debt) > 0 else 0,
                    'at': rev / ta if ta else 0,
                    'gm': 0.38 if key not in ('fs','bpo','treas') else 0.45,
                    'eb_pct': eb / rev if rev else 0,
                    'nm': pat / rev if rev else 0,
                    'liq': cash / st_debt if st_debt else 0,
                }[rkey]
                cell = ws.cell(row=row, column=3 + yi, value=round(v, 4 if fmt == PCT_FMT else 2))
                cell.number_format = fmt
                cell.font = BODY_FONT
                # Critical ratios — tag Red if covenant likely tripped
                if rkey == 'nd_eb' and v > 3.5:
                    cell.fill = PatternFill('solid', fgColor='FFC7CE')
                if rkey == 'int_cov' and v < 3.0:
                    cell.fill = PatternFill('solid', fgColor='FFC7CE')
                if rkey == 'liq' and v < 1.0:
                    cell.fill = PatternFill('solid', fgColor='FFEB9C')
            row += 1
        row += 1  # gap row between entities

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    for col in range(3, 8):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.freeze_panes = 'C5'


# ── Sheet 6: Variance Drill-Down (FY2025 Q4) ───────────────────────────────
def build_variance_drilldown(ws):
    ws.title = 'Variance Drill-Down'
    ws['A1'] = 'FY2025 Q4 Variance Drill-Down — 11 Divisions × Top 8 Drivers'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:H1')
    ws['A2'] = 'Volume / Price / Mix / Cost / FX impact in MYR M | Source: Division CFOs and Group FP&A'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:H2')

    headers = ['Division', 'Driver', 'Volume', 'Price', 'Mix', 'Cost', 'FX', 'Total EBITDA Impact']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_header(c)

    DRIVERS = {
        'chem':   ['Lower olefins demand from China','Specialty chemicals customer cancellations','Petrochemical feedstock spike','Energy & utilities cost','Sales mix shift to commodity grade','MYR/USD adverse FX','Plant utilisation drop at Penang','Logistics & freight'],
        'agri':   ['CPO price decline 18%','Indonesian export levy increase','Yield decline at Sumatra estates','Replanting capex acceleration','Labour cost inflation','IDR weakening vs USD','Manuring cost increase','Refinery throughput shortfall'],
        'mfg':    ['Auto OEM volume softness','Steel and aluminium cost inflation','Penang plant downtime - 11 days','Semiconductor packaging margin compression','Aerospace contract delay','Energy cost spike','Logistics & freight','Customer working-capital terms'],
        'fs':     ['Improved investment income','Margin lending growth','Underwriting fees beat','Credit cost release','Wealth AUM growth','Performance fee recognition','Cost discipline','Tax efficiency'],
        'prop':   ['Mid-Valley project completion delay','Penang Tower writedown','Inventory carrying cost','Sales velocity slow','Klang Valley unsold stock','Land bank revaluation','Construction cost inflation','Marketing & sales cost surge'],
        'bpo':    ['New US client wins','Pricing escalators','AI-led productivity gains','Bench utilisation improvement','Currency tailwind','Premises cost optimisation','Health insurance benefit','Revenue mix shift to higher-value'],
        'trade':  ['Crude oil curve flattening','Carry trade losses','Counterparty default - Hindustan Bulk','Inventory mark-to-market','Storage and freight cost','LME margin call event','FX hedging slippage','Trade finance cost increase'],
        'pharma': ['Generics tender wins (Indonesia)','API cost inflation','New SKU contribution','Active ingredient procurement','BPOM regulatory delay','Distribution cost optimisation','Currency hedging gain','Channel restock'],
        'hc':     ['Patient volume decline post-pandemic','Specialist consultant cost','Subang facility under-utilisation','Insurance reimbursement cuts','Pharmacy margin compression','Equipment lease expiry','Bad debt provisioning','Staff cost inflation'],
        'retail': ['Footfall decline at flagship stores','Online cannibalisation','Inventory write-down','Lease cost inflation','Marketing campaign over-spend','Buying margin compression','Loyalty programme cost','Digital transformation cost'],
        'treas':  ['Investment portfolio mark-down','FX translation losses','Interest rate hedge slippage','Group cost allocation reset','Deposit yield drop','Fund management fees','Custody fees','Audit & regulatory cost'],
    }

    # Generate plausible distributions of impact by driver
    import random
    rng = random.Random(42)
    row = 6
    for did, dname, _, _, _ in DIVISIONS:
        ws.cell(row=row, column=1, value=dname).font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        for col in range(2, 9):
            ws.cell(row=row, column=col).fill = SUBHEADER_FILL
        row += 1

        # Total EBITDA variance for this division
        total_var = (FY25[did][0] * FY25[did][2] - FY25[did][1] * FY25[did][3]) * -1  # Bud - Act, flipped to negative
        # Actually let me keep it as actual - budget (so negative = miss)
        bud_eb = fy25_eb(did, 'bud_eb')
        act_eb = fy25_eb(did, 'act_eb')
        total_eb_var = act_eb - bud_eb

        for di, driver in enumerate(DRIVERS[did]):
            ws.cell(row=row, column=1, value='').font = BODY_FONT
            ws.cell(row=row, column=2, value=driver).font = BODY_FONT
            # Distribute total var across 8 drivers with random weights
            weights = [rng.uniform(0.5, 2.5) for _ in range(8)]
            wsum = sum(weights)
            weight = weights[di] / wsum
            driver_impact = total_eb_var * weight  # MYR M
            # Split each driver's impact across 5 attribution columns
            attribution = [rng.uniform(0, 1) for _ in range(5)]
            asum = sum(attribution)
            attr_norm = [a / asum * driver_impact for a in attribution]
            for j, val in enumerate(attr_norm):
                cell = ws.cell(row=row, column=3 + j, value=round(val, 1))
                cell.number_format = NUM_FMT
                cell.font = BODY_FONT
                if val < -10:
                    cell.font = Font(color='C00000', size=10, name='Aptos')
            cell = ws.cell(row=row, column=8, value=round(driver_impact, 1))
            cell.number_format = NUM_FMT
            cell.font = Font(bold=True, color='C00000' if driver_impact < 0 else '00B050', size=10, name='Aptos')
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = BORDER
            row += 1
        # Division subtotal
        ws.cell(row=row, column=1, value='').font = TOTAL_FONT
        ws.cell(row=row, column=2, value=f'{dname} Subtotal').font = TOTAL_FONT
        ws.cell(row=row, column=2).fill = TOTAL_FILL
        for col_idx in range(3, 9):
            ws.cell(row=row, column=col_idx).fill = TOTAL_FILL
        cell = ws.cell(row=row, column=8, value=round(total_eb_var, 1))
        cell.number_format = NUM_FMT
        cell.font = Font(bold=True, color='C00000' if total_eb_var < 0 else '00B050', size=11, name='Aptos')
        cell.fill = TOTAL_FILL
        row += 1
        row += 1  # gap row

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 50
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = 'B5'


def main(out_path):
    wb = Workbook()
    # Remove default sheet, then add 6 named sheets
    wb.remove(wb.active)
    for builder in (build_pl_summary, build_revenue_bridge, build_quarterly_tracker,
                    build_balance_sheet, build_ratios, build_variance_drilldown):
        ws = wb.create_sheet()
        builder(ws)
    wb.save(out_path)
    print(f'Wrote {out_path}')


if __name__ == '__main__':
    out = sys.argv[1] if len(sys.argv) > 1 else 'files/01_Zava_Group_Financial_Performance.xlsx'
    main(out)
