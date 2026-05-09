"""Generates 05_Zava_Lender_Covenant_Tracker.xlsx — 4 sheets, 18 lenders.

Sheets:
  1. Lender Register — 18 lenders × 8 attributes
  2. Covenant Matrix — 18 facilities × 6 covenants × 4 status fields
  3. Covenant Headroom Trend — 6 key facilities × 8 quarters × 4 covenants
  4. Waiver/Amendment History — 12 historical events × 7 attributes
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from gen_ref_files.common import LENDERS

NAVY_FILL = PatternFill('solid', fgColor='1F4E79')
SUB_FILL = PatternFill('solid', fgColor='2E75B6')
TOTAL_FILL = PatternFill('solid', fgColor='D9E1F2')
RED_FILL = PatternFill('solid', fgColor='C00000')
AMBER_FILL = PatternFill('solid', fgColor='FFC000')
GREEN_FILL = PatternFill('solid', fgColor='00B050')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Aptos')
SUB_FONT = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
BODY_FONT = Font(size=10, name='Aptos')
TOTAL_FONT = Font(bold=True, size=10, name='Aptos')
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0;[Red](#,##0);-'
PCT_FMT = '0.0%;[Red]-0.0%;-'
RATIO_FMT = '0.00'


def style_h(c, fill=NAVY_FILL, font=HEADER_FONT):
    c.fill = fill
    c.font = font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER


def status_pill(cell, status):
    if status == 'red':
        cell.fill = RED_FILL
        cell.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
    elif status == 'amber':
        cell.fill = AMBER_FILL
        cell.font = Font(bold=True, color='000000', size=10, name='Aptos')
    else:
        cell.fill = GREEN_FILL
        cell.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')


# ── Sheet 1: Lender Register ───────────────────────────────────────────────
def s1_register(ws):
    ws.title = 'Lender Register'
    ws['A1'] = 'Zava Group Lender Register — 18 Active Facilities'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:I1')
    ws['A2'] = 'Total committed: as below | Source: Group Treasury | Last refresh: FY2025 Q4'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:I2')

    headers = ['Code', 'Lender / Instrument', 'Country', 'Facility Type', 'Committed (MYR M)', 'Drawn (MYR M)', 'Undrawn (MYR M)', 'Maturity', 'Rate Type', 'Security']
    for i, h in enumerate(headers, 1):
        style_h(ws.cell(row=4, column=i, value=h))

    row = 5
    total_committed = 0
    total_drawn = 0
    for code, lender, country, ftype, committed, drawn, maturity, rate_type, security in LENDERS:
        ws.cell(row=row, column=1, value=code).font = Font(bold=True, size=10, name='Aptos')
        ws.cell(row=row, column=2, value=lender).font = BODY_FONT
        ws.cell(row=row, column=3, value=country).font = BODY_FONT
        ws.cell(row=row, column=4, value=ftype).font = BODY_FONT
        ws.cell(row=row, column=5, value=committed).number_format = NUM_FMT
        ws.cell(row=row, column=6, value=drawn).number_format = NUM_FMT
        ws.cell(row=row, column=7, value=committed - drawn).number_format = NUM_FMT
        ws.cell(row=row, column=8, value=maturity).font = BODY_FONT
        ws.cell(row=row, column=9, value=rate_type).font = BODY_FONT
        ws.cell(row=row, column=10, value=security).font = BODY_FONT
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = BORDER
        total_committed += committed
        total_drawn += drawn
        row += 1

    ws.cell(row=row, column=1, value='').fill = TOTAL_FILL
    ws.cell(row=row, column=2, value='TOTAL ZAVA GROUP DEBT').font = TOTAL_FONT
    ws.cell(row=row, column=2).fill = TOTAL_FILL
    ws.cell(row=row, column=3, value='').fill = TOTAL_FILL
    ws.cell(row=row, column=4, value='').fill = TOTAL_FILL
    for col, val in [(5, total_committed), (6, total_drawn), (7, total_committed - total_drawn)]:
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = NUM_FMT
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
    for col in range(8, 11):
        ws.cell(row=row, column=col).fill = TOTAL_FILL

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 22
    for col in ['E', 'F', 'G']:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 22
    ws.freeze_panes = 'B5'


# ── Sheet 2: Covenant Matrix ──────────────────────────────────────────────
def s2_covenants(ws):
    ws.title = 'Covenant Matrix'
    ws['A1'] = 'Covenant Matrix — All 18 Facilities × 6 Covenants × 4 Fields'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:H1')
    ws['A2'] = 'Status: Green=>20% headroom, Amber=10-20%, Red=<10% or breach. As of FY2025 Q4.'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:H2')

    headers = ['Lender Code', 'Lender / Instrument', 'Covenant', 'Limit', 'Current', 'Headroom', 'Headroom %', 'Status']
    for i, h in enumerate(headers, 1):
        style_h(ws.cell(row=4, column=i, value=h))

    # FY25 actuals: Net Debt/EBITDA = 3.1, Interest Cover = 4.2, Current = 1.3, Min Liq cushion ~ 1.4× ST debt
    COVENANT_FORMULAS = [
        # (covenant name, limit, current, fmt, lower_better=True)
        ('Net Debt / EBITDA', 3.5, 3.1, RATIO_FMT, True),
        ('Interest Cover (×)', 3.0, 4.2, RATIO_FMT, False),
        ('Current Ratio (×)', 1.0, 1.3, RATIO_FMT, False),
        ('Min Liquidity (× ST Debt)', 1.0, 1.4, RATIO_FMT, False),
        ('Min Equity (MYR M)', 18000, 22400, NUM_FMT, False),
        ('MAC clause', 'No MAC', 'Compliant', None, None),
    ]

    row = 5
    for code, lender, country, ftype, committed, drawn, maturity, rate_type, security in LENDERS:
        c = ws.cell(row=row, column=1, value=code)
        c.font = SUB_FONT
        c.fill = SUB_FILL
        ws.cell(row=row, column=2, value=lender).font = SUB_FONT
        ws.cell(row=row, column=2).fill = SUB_FILL
        for col in range(3, 9):
            ws.cell(row=row, column=col).fill = SUB_FILL
        row += 1
        for cov_name, limit, current, fmt, lower_better in COVENANT_FORMULAS:
            ws.cell(row=row, column=1, value='').font = BODY_FONT
            ws.cell(row=row, column=2, value='').font = BODY_FONT
            ws.cell(row=row, column=3, value=cov_name).font = BODY_FONT
            if fmt is None:
                # MAC clause
                ws.cell(row=row, column=4, value=str(limit)).font = BODY_FONT
                ws.cell(row=row, column=5, value=str(current)).font = BODY_FONT
                ws.cell(row=row, column=6, value='—').font = BODY_FONT
                ws.cell(row=row, column=7, value='—').font = BODY_FONT
                sc = ws.cell(row=row, column=8, value='GREEN')
                status_pill(sc, 'green')
                sc.alignment = Alignment(horizontal='center')
            else:
                if lower_better:
                    headroom = limit - current
                    headroom_pct = headroom / limit if limit else 0
                else:
                    headroom = current - limit
                    headroom_pct = headroom / limit if limit else 0
                ws.cell(row=row, column=4, value=limit).number_format = fmt
                ws.cell(row=row, column=5, value=current).number_format = fmt
                ws.cell(row=row, column=6, value=round(headroom, 2 if fmt == RATIO_FMT else 0)).number_format = fmt
                ws.cell(row=row, column=7, value=headroom_pct).number_format = PCT_FMT
                # Status
                if headroom_pct < 0:
                    status = 'red'
                    label = 'BREACH'
                elif headroom_pct < 0.10:
                    status = 'red'
                    label = 'TIGHT'
                elif headroom_pct < 0.20:
                    status = 'amber'
                    label = 'AMBER'
                else:
                    status = 'green'
                    label = 'GREEN'
                sc = ws.cell(row=row, column=8, value=label)
                status_pill(sc, status)
                sc.alignment = Alignment(horizontal='center')
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = BORDER
            row += 1

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 26
    for col in ['D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['H'].width = 10
    ws.freeze_panes = 'C5'


# ── Sheet 3: Covenant Headroom Trend ──────────────────────────────────────
def s3_trend(ws):
    ws.title = 'Headroom Trend'
    ws['A1'] = 'Covenant Headroom Trend — 6 Key Facilities × 8 Quarters × 4 Covenants'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:K1')
    ws['A2'] = 'Q1 FY24 to Q4 FY25 | Net Debt/EBITDA, Interest Cover, Current Ratio, Min Liquidity'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:K2')

    quarters = ['Q1 FY24', 'Q2 FY24', 'Q3 FY24', 'Q4 FY24', 'Q1 FY25', 'Q2 FY25', 'Q3 FY25', 'Q4 FY25']
    headers = ['Facility', 'Covenant'] + quarters + ['Limit']
    for i, h in enumerate(headers, 1):
        style_h(ws.cell(row=4, column=i, value=h))

    # 6 key facilities (use first 6 from LENDERS that are large)
    KEY_FACILITIES = [LENDERS[i] for i in (0, 1, 2, 3, 5, 6)]  # cherry-pick the largest
    # Sample trends — Q4 FY25 reflects current actual; backfill plausibly
    TRENDS = {
        'NetDebt': {
            'limit': 3.5, 'lower_better': True,
            'values': [2.1, 2.2, 2.3, 2.4, 2.5, 2.7, 2.9, 3.1],
        },
        'IntCover': {
            'limit': 3.0, 'lower_better': False,
            'values': [5.8, 5.6, 5.4, 5.1, 4.9, 4.7, 4.5, 4.2],
        },
        'CurrRatio': {
            'limit': 1.0, 'lower_better': False,
            'values': [1.5, 1.5, 1.4, 1.4, 1.4, 1.3, 1.3, 1.3],
        },
        'Liquidity': {
            'limit': 1.0, 'lower_better': False,
            'values': [1.8, 1.7, 1.7, 1.6, 1.5, 1.5, 1.4, 1.4],
        },
    }

    row = 5
    for code, lender, country, ftype, committed, drawn, maturity, rate_type, security in KEY_FACILITIES:
        # Sub-header for each facility
        c = ws.cell(row=row, column=1, value=f'{code}: {lender[:32]}')
        c.font = SUB_FONT
        c.fill = SUB_FILL
        for col in range(2, 12):
            ws.cell(row=row, column=col).fill = SUB_FILL
        row += 1
        for cov_name, cov_data in [
            ('Net Debt / EBITDA', TRENDS['NetDebt']),
            ('Interest Cover (×)', TRENDS['IntCover']),
            ('Current Ratio (×)', TRENDS['CurrRatio']),
            ('Min Liquidity (× ST Debt)', TRENDS['Liquidity']),
        ]:
            ws.cell(row=row, column=1, value='').font = BODY_FONT
            ws.cell(row=row, column=2, value=cov_name).font = BODY_FONT
            for qi, val in enumerate(cov_data['values']):
                # Add small per-facility variation
                v = val + ((hash(code + str(qi)) % 11) - 5) / 100
                cell = ws.cell(row=row, column=3 + qi, value=round(v, 2))
                cell.number_format = RATIO_FMT
                cell.font = BODY_FONT
                # Highlight breach or tight
                if cov_data['lower_better']:
                    if v > cov_data['limit']:
                        cell.fill = RED_FILL
                        cell.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
                    elif v > cov_data['limit'] * 0.9:
                        cell.fill = AMBER_FILL
                else:
                    if v < cov_data['limit']:
                        cell.fill = RED_FILL
                        cell.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
                    elif v < cov_data['limit'] * 1.1:
                        cell.fill = AMBER_FILL
            ws.cell(row=row, column=11, value=cov_data['limit']).number_format = RATIO_FMT
            ws.cell(row=row, column=11).font = TOTAL_FONT
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = BORDER
            row += 1
        row += 1  # gap

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 24
    for col in range(3, 12):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.freeze_panes = 'C5'


# ── Sheet 4: Waiver/Amendment History ──────────────────────────────────────
def s4_history(ws):
    ws.title = 'Waiver History'
    ws['A1'] = 'Waiver, Amendment & Standstill History — Last 12 Material Events'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Aptos')
    ws.merge_cells('A1:G1')
    ws['A2'] = 'Source: Group Treasury Legal | All material events disclosed to Audit & Risk Committee'
    ws['A2'].font = Font(italic=True, size=9, color='595959', name='Aptos')
    ws.merge_cells('A2:G2')

    headers = ['Date', 'Lender / Code', 'Event Type', 'Trigger', 'Action Taken', 'Outcome', 'Status']
    for i, h in enumerate(headers, 1):
        style_h(ws.cell(row=4, column=i, value=h))

    EVENTS = [
        ('Mar 2021', 'L01 — Maybank RCF', 'Amendment', 'Sustainability-linked KPI insertion', 'Pricing margin tied to Scope 1+2 reduction (12.5 bps step-up)', 'Executed; FY2024 KPI met', 'Closed'),
        ('Aug 2021', 'L05 — CIMB Bilateral', 'Waiver', 'Negative pledge covenant clarification', 'Carve-out for project finance subsidiaries', 'Granted', 'Closed'),
        ('Feb 2022', 'L08 — RHB Sukuk', 'Amendment', 'Refinancing — extension of tenor', 'Tenor extended by 3 years; spread renegotiated', 'Approved', 'Closed'),
        ('Jul 2022', 'L11 — DBS USD Term Loan', 'Standstill', 'Temporary FX-driven covenant tightness', 'Standstill granted for 6 months', 'Resolved by FY-end', 'Closed'),
        ('Nov 2022', 'L14 — OCBC SGD MTN', 'Amendment', 'Cross-default threshold harmonisation', 'Aligned across all SGD instruments', 'Approved', 'Closed'),
        ('Mar 2023', 'L02 — HSBC Term Loan', 'Amendment', 'Sustainability-linked margin step', 'Pricing margin tied to renewable electricity %', 'Executed; FY2024 KPI met', 'Closed'),
        ('Sep 2023', 'L09 — Bank Mandiri Sukuk', 'Amendment', 'Permitted-acquisition basket increase', 'Basket increased from MYR 500M to MYR 1B', 'Approved', 'Closed'),
        ('Feb 2024', 'L13 — UOB MYR Term Loan', 'Waiver', 'One-time exception for Penang Tower writedown', 'Net Debt/EBITDA waiver Q1 FY2024 only', 'Granted', 'Closed'),
        ('May 2024', 'L17 — JP Morgan USD MTN', 'Information', 'Material non-public information sharing protocol update', 'Updated information-sharing agreement', 'Executed', 'Closed'),
        ('Sep 2024', 'L03 — Public Bank RCF', 'Amendment', 'Hindustan Bulk counterparty event', 'Treasury policy strengthened; trade-credit insurance mandate', 'Executed', 'Closed'),
        ('Nov 2024', 'L06 — AmBank Bilateral', 'Discussion', 'Anticipated Q4 FY2025 covenant tightness', 'Pre-emptive engagement; quarterly check-ins agreed', 'Ongoing', 'Active'),
        ('Dec 2024', 'L01-L18 — All Lenders', 'Notification', 'FY2025 Q4 Net Debt/EBITDA expected at 3.1×', 'All lenders notified per CIA; quarterly call scheduled Q1 FY2026', 'Notified', 'Active'),
    ]

    row = 5
    for date, lender, etype, trigger, action, outcome, status in EVENTS:
        ws.cell(row=row, column=1, value=date).font = BODY_FONT
        ws.cell(row=row, column=2, value=lender).font = BODY_FONT
        ws.cell(row=row, column=3, value=etype).font = BODY_FONT
        ws.cell(row=row, column=4, value=trigger).font = BODY_FONT
        ws.cell(row=row, column=5, value=action).font = BODY_FONT
        ws.cell(row=row, column=6, value=outcome).font = BODY_FONT
        sc = ws.cell(row=row, column=7, value=status)
        if status == 'Closed':
            sc.fill = GREEN_FILL
            sc.font = Font(bold=True, color='FFFFFF', size=10, name='Aptos')
        else:
            sc.fill = AMBER_FILL
            sc.font = Font(bold=True, color='000000', size=10, name='Aptos')
        sc.alignment = Alignment(horizontal='center')
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 38
    ws.column_dimensions['E'].width = 42
    ws.column_dimensions['F'].width = 32
    ws.column_dimensions['G'].width = 10
    ws.freeze_panes = 'B5'


def main(out_path):
    wb = Workbook()
    wb.remove(wb.active)
    for builder in (s1_register, s2_covenants, s3_trend, s4_history):
        ws = wb.create_sheet()
        builder(ws)
    wb.save(out_path)
    print(f'Wrote {out_path}')


if __name__ == '__main__':
    out = sys.argv[1] if len(sys.argv) > 1 else 'files/05_Zava_Lender_Covenant_Tracker.xlsx'
    main(out)
